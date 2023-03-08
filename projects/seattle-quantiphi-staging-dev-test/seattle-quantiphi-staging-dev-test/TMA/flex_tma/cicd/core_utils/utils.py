# Run these commands first for installs
#    >>> !pip install openslide-python
#    >>> !pip install opencv-python
#    >>> !sudo apt-get install -y openslide-tools
#    >>> !sudo apt-get update
#    >>> !sudo apt-get install ffmpeg libsm6 libxext6  -y

# Run these commands next to get the data
#    >>> !mkdir -p ./data
#    >>> !gsutil -m cp -r gs://ihc_dataset/quantiphi_testing/* ./data

# Data Science Imports
import pandas as pd
import numpy as np
import tensorflow as tf
import random

# Microscopy Related Imports
import openslide
import tifffile

# Cloud Related Imports
from google.protobuf.struct_pb2 import Value
from google.protobuf import json_format
from google.cloud import aiplatform
from google.cloud import storage

# Visualization Imports
from matplotlib.colors import ListedColormap
from matplotlib.patches import Patch
import matplotlib.pyplot as plt
from tqdm.notebook import tqdm
import seaborn as sns
from PIL import Image
import cv2
import skimage; import skimage.color; import skimage.filters;

import apache_beam as beam
from apache_beam import pvalue

# Built In Imports
import base64
import glob
import os
import gc
import sys
import logging
import shutil
import psutil

from openpyxl import load_workbook
from google.cloud import bigquery


def load_file_gcs(gcs_path, destination_path):
    """
    loads  object in Google Cloud Storage bucket to local
    
    Args:
        gcs_path(str): path to file in gcs bucket
        destination_path(str): path to local directory to store the file
    
    Returns:
        local file path for the file
    """
    bucket_name = gcs_path.split('/')[2]       #extract bucket name
    source_blob_path = '/'.join(gcs_path.split('/')[3:])       #extract folder structure
    filename = source_blob_path.split('/')[-1]      #extract file name
    storage_client = storage.Client("[ihc_qc_sanbox]")
        # Create a bucket object for our bucket
    bucket = storage_client.get_bucket(bucket_name)
        # Create a blob object from the filepath
    blob = bucket.blob(source_blob_path)
        # Download the file to a destination
    blob.download_to_filename(destination_path+filename)
    logging.getLogger().setLevel(logging.INFO)
    logging.info("Slide Loaded from GCS bucket as : %s",str(destination_path+filename))
    
    return str(destination_path+filename)

def predict_image_object_detection_sample(
    items,
    project: str,
    endpoint_id: str,
    C_THRESH,
    config,
    job_name,
    location: str = "us-central1",
    api_endpoint: str = "us-central1-prediction-aiplatform.googleapis.com",
    
    
):
    """ Get predictions from Google AIPlatform Endpoint """
    client_options = {"api_endpoint": api_endpoint}
        
    # Initialize client that will be used to create and send requests.
    # This client only needs to be created once, and can be reused for multiple requests.
    client = aiplatform.gapic.PredictionServiceClient(client_options=client_options)
    
    gcs_path,image_class,filename,local_svs_path = items
    with open(filename, "rb") as f:
        file_content = f.read()

    # The format of each instance should conform to the deployed model's prediction input schema.
    encoded_content = base64.b64encode(file_content).decode("utf-8")
    instance_dict = {"content": encoded_content}
    
    instance = json_format.ParseDict(instance_dict, Value())
    instances = [instance]
    
    # See gs://google-cloud-aiplatform/schema/predict/params/image_object_detection_1.0.0.yaml for the format of the parameters.
    
    parameters_dict = {"confidenceThreshold": C_THRESH, "maxPredictions": 200}
    parameters = json_format.ParseDict(parameters_dict, Value())
    endpoint = client.endpoint_path(
        project=project, location=location, endpoint=endpoint_id
    )
    response = client.predict(
        endpoint=endpoint, instances=instances, parameters=parameters
    )
    logging.getLogger().setLevel(logging.INFO)
    deployed_model_id = response.deployed_model_id
    logging.info(" deployed_model_id:  "+ str(deployed_model_id))
    predictions = response.predictions
    filename = filename.replace("Gray","Original")
    if (np.array(predictions[0]["bboxes"]).size == 0):
        logging.exception("Model response is blank")        
        yield pvalue.TaggedOutput("Error_Tag",(gcs_path.split(',')[0],config["function_config"]["Insert_Error_BQ"]["dataset_id"],config["function_config"]["Insert_Error_BQ"]["table_id"],job_name,'Prediction',"Object detection model response is blank"))
        
    else:
        yield pvalue.TaggedOutput("Process_Tag",(gcs_path, np.array(predictions[0]["confidences"]), np.array(predictions[0]["bboxes"]), filename, local_svs_path,image_class))

def get_bbox_from_svs(path, bbox_frac, resize_to=2048):
    """ Simple function to open an SVS image and downscale
    
    Args:
        path (str): Path to the svs file to be returned
        bbox_frac (list of floats): The bounding box coordinates
            as fractional values (xmin, ymin, xmax, ymax)
        resize_to (int, optional): The value to resize the image to (for model inference)
    
    Return:
        PIL Image of the desired core/bbox
    """
    
    # Open the image
    slide = openslide.OpenSlide(path)
    
    # Get dimensions
    slide_w, slide_h = slide.dimensions
    
    # Get bbox coordinates in native scale
    xmin,  ymin  = int(bbox_frac[0]*slide_w), int(bbox_frac[1]*slide_h)
    box_w, box_h = int((bbox_frac[2]-bbox_frac[0])*slide_w), int((bbox_frac[3]-bbox_frac[1])*slide_h)

    # Cut out the bbox
    bbox_cut_raw = slide.read_region(location=(xmin,ymin), level=0, size=(box_w, box_h))
    return Image.fromarray(cv2.resize(np.asarray(bbox_cut_raw), (resize_to,resize_to), cv2.INTER_AREA))


def save_cores_to_disk(bb_lbl_map, path_to_svs, output_dir=None, save_as_rgb=True):
    """ Function to save the cores in outut directory """
    # Add required folder(s) and define paths
    tma_subdir, file_name = demo_path.rsplit(".", 1)[0].rsplit("/", 2)[1:]
    output_dir = os.path.join(output_dir, tma_subdir, file_name)
    if not os.path.isdir(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    for row_ltr, bbox_list in tqdm(demo_bb_lbl_map.items(), total=len(demo_bb_lbl_map)):
        for row_idx, bbox in tqdm(enumerate(bbox_list), total=len(bbox_list)):
            
            individual_core_fname = os.path.join(output_dir, f"{file_name}__{row_ltr}{row_idx:02}.jpg")
            core_img = get_bbox_from_svs(demo_path, bbox)
            if save_as_rgb:
                core_img.convert("RGB").save(individual_core_fname.replace(".jpg", "_rgb.jpg"), quality=100)
            else:
                core_img.convert("L").save(individual_core_fname.replace(".jpg", "_gray.jpg"), quality=100)
    
    
def grab_contours(cnts):
    """ Return filtered contours -- from imutils library """
    # if the length the contours tuple returned by cv2.findContours
    # is '2' then we are using either OpenCV v2.4, v4-beta, or
    # v4-official
    if len(cnts) == 2:
        cnts = cnts[0]

    # if the length of the contours tuple is '3' then we are using
    # either OpenCV v3, v4-pre, or v4-alpha
    elif len(cnts) == 3:
        cnts = cnts[1]

    # otherwise OpenCV has changed their cv2.findContours return
    # signature yet again and I have no idea WTH is going on
    else:
        raise Exception(("Contours tuple must have length 2 or 3, "
            "otherwise OpenCV changed their cv2.findContours return "
            "signature yet again. Refer to OpenCV's documentation "
            "in that case"))

    # return the actual contours array
    return cnts


def get_wsi(file_path, cloud_dir="", save_to_disk=True, save_rgb_to_disk=False):
    """ Get WSI from cloud storage or local and save original/downsampled/rgb to disk 
    
    Args:
        file_path (str): Path to the slide on the local
        cloud_dir (str): Subfolder containing the image on the cloud to download
        save_to_disk (bool, optional): Whether to save the grayscale image
            to the local harddrive
        save_rgb_to_disk (bool, optional): Whether to save the RGB image
            to the local harddrive
    
    Returns: None; or, if `save_to_disk=False`, a numpy array containing the
        downscaled image.
    
    """
    
    # Get filename from filepath
    filename = file_path.rsplit("/", 1)[1]
    
    if not os.path.isfile(file_path):
        # Verbose What We're Doing
        logging.getLogger().setLevel(logging.INFO)
        logging.info("Downloading : %s",str(file_path))
    
        # Figure out which of the 4 buckets the file is in and get the respective blob
        if not cloud_dir:
            dir_blob_tmp = [
                (tma_dir, bucket.get_blob(os.path.join(tma_dir.rsplit("/", 1)[1], filename))) \
                for tma_dir in TMA_DIRS
            ]
            cloud_dir, blob = [dir_blob for dir_blob in dir_blob_tmp if dir_blob[1] is not None][0]
    
        # Create the destination filepath
        dest_filepath = os.path.join(cloud_dir, filename)
    
        # Download the file from GCS
        blob.download_to_filename(dest_filepath)
    else:
        cloud_dir = file_path.rsplit("/", 2)[1] # subdirectory
        dest_filepath = file_path # where it's located on our local
        logging.getLogger().setLevel(logging.INFO)
        logging.info("--> Already Downloaded")

    # Read the image dimensions
    logging.info("\n... Processing %s ...",filename)
    wsi, dwn_sample, dwn_dimensions = open_svs(file_path)

    # Verbose Old Dimensions
    logging.info("\t--> Original Slide Dimensions: " \
          f"{tuple([round(x*dwn_sample,0) for x in dwn_dimensions])}")
    
    # Verbose New Dimensions
    logging.info("\t--> New Slide Dimensions: " \
          f"{dwn_dimensions}")

    # Save RGB to disk
    if save_rgb_to_disk:
        rgb_png_path = dest_filepath.replace(f".{filename.rsplit('.', 1)[1]}", 
                                             f"_{dwn_sample:.3f}_{dwn_dimensions[0]}_{dwn_dimensions[1]}.png")\
                                    .replace("svs_files", "rgb_downsampled")
        
        # Check if it exists and make if necessary
        if not os.path.isdir(rgb_png_path.rsplit("/", 1)[0]):
            os.makedirs(rgb_png_path.rsplit("/", 1)[0], exist_ok=True)
        
        wsi.convert("RGB").save(rgb_png_path, quality=100)

        logging.info(f"\t--> Saved RGB Image as PNG To Disk: {rgb_png_path}")
    
    # Convert To Graysale
    wsi = wsi.convert("L")
    logging.info(f"\t--> Resized Slide Dimensions: {dwn_dimensions}")

    # Save Grayscale to disk
    if save_to_disk:
        gray_png_path = dest_filepath.replace(f".{filename.rsplit('.', 1)[1]}", 
                                              f"_{dwn_sample:.3f}_{dwn_dimensions[0]}_{dwn_dimensions[1]}.png")\
                                     .replace("svs_files", "downsampled")

        # Check if it exists and make if necessary
        if not os.path.isdir(gray_png_path.rsplit("/", 1)[0]):
            os.makedirs(gray_png_path.rsplit("/", 1)[0], exist_ok=True)
        
        wsi.save(gray_png_path, quality=100)
        logging.info(f"\t--> Saved as Grayscale PNG To Disk: {gray_png_path}")
        return gray_png_path
    else:
        logging.info(f"\t--> Returning Image As Numpy Array")
        return np.asarray(wsi), dwn_sample, dwn_dimensions
    

def get_bb_iou(box_a, box_b):
    """ Function to compute boundary box"""
    
    # determine the (x,y)-coordinates of the intersection rectangle
    x_a = max(box_a[0], box_b[0])
    y_a = max(box_a[1], box_b[1])
    x_b = min(box_a[2], box_b[2])
    y_b = min(box_a[3], box_b[3])

    # compute the area of intersection rectangle
    #     - do not include a +1 in these area calculations
    inter_area = max(0, x_b - x_a) * max(0, y_b - y_a)

    # compute the area of both the prediction and ground-truth rectangles
    #     - do not include a +1 in these area calculations
    box_a_area = (box_a[2] - box_a[0]) * (box_a[3] - box_a[1])
    box_b_area = (box_b[2] - box_b[0]) * (box_b[3] - box_b[1])
    
    if box_a_area > box_b_area:
        max_box_area = 'box_a'
    else:
        max_box_area = 'box_b'
        

    # compute the intersection over union by taking the intersection
    # area and dividing it by the sum of prediction + ground-truth
    # areas - the interesection area
    iou = inter_area / float(box_a_area + box_b_area - inter_area)

    # return the intersection over union value and box with max area
    return iou, max_box_area


def bbox_suppression(bboxes, centre_points, overlap_thresh=0.2):
    """ Function to suppress the boundary box """
    
    # if there are no boxes, return an empty list
    if len(bboxes) == 0:
        return []
    
    # initialize the list of picked indexes	
    indices_to_remove = []
    for i in range(len(bboxes)):
        for j in range(i+1, len(bboxes)):
            ith_bbox = bboxes[i]
            jth_bbox = bboxes[j]
            
            iou, max_box_area_str = get_bb_iou(ith_bbox, jth_bbox)
            if iou > overlap_thresh:
                
                # ith and jth box area - compare and return the winner
                if max_box_area_str == 'box_a':
                    indices_to_remove.append(i)
                else:
                    indices_to_remove.append(j)
                    
    bboxes = np.delete(bboxes, indices_to_remove, axis=0)
    centre_points = np.delete(centre_points, indices_to_remove, axis=0)
    
    return bboxes, centre_points

def compute_bb_rows_1(bbs, cps, make_flexible=True):
    """ Get the bounding boxes by row """
    def __get_top_band(boxes, expansion_factor=0.01):
        """ Find the top band of bounding boxes """
        top_row = boxes[np.argmin(boxes[:, 1])][[1,3]]
        if expansion_factor:
            return np.array((top_row[0]*(1-expansion_factor), top_row[1]*(1+expansion_factor)))
        
    bb_rows = []
    while len(bbs)!=0:
        top_row_band = __get_top_band(bbs)
        top_row_idxs = np.where(np.less(cps[:, 1], top_row_band[1]))[0]
        top_row_bbs = bbs[top_row_idxs, :]

        if make_flexible:
            ymax = top_row_bbs[np.argmax(top_row_bbs[:, 3])][3]
            o_ymax = top_row_bbs[np.argmin(top_row_bbs[:, 3])][3]
            ymax = (ymax+o_ymax)/2
            top_row_idxs = np.where(np.less(cps[:, 1], ymax))[0]
            top_row_bbs = bbs[top_row_idxs, :]

        # Sort left-to-right
        sorted_top_row_bbs = top_row_bbs[np.argsort(top_row_bbs[:, 0])]

        # Append sorted list to master list
        bb_rows.append(sorted_top_row_bbs)
        
        # Remove from original array
        bbs = np.delete(bbs, top_row_idxs, axis=0)
        cps = np.delete(cps, top_row_idxs, axis=0)
        
    return bb_rows


def closest_cp(cp, cps, max_dy, x_step=1, y_step=5):
    """ Algorithmic way to find the closest Centre Point 
    
    This uses a modified version of Euclidean Distance.
    In the new coordinate system x is non-continuous 
        - negative jumps to very large positive
    and steps in the y-direction are much more costly
    
    Args:
        cp (nd.Array): Numpy array containing the centre-point 
            we want to use to calculate distance based on
        cps (nd.Array): Other centre-points for which we will calculate
            the closest centre-point to the original `cp` centre point.
        max_dy (float): Maximum distance the next centre-point can exist
            in the y-direction
        x_step (int, optional): The respective distance of one horizontal step in
            our modified coordinate system
        y_step (int, optional): The respective distance of one vertical step in
            our modified coordinate system
        """
    # Find the x distances between cp and all cps
    delta_x = (cps[..., 0] - cp[0])
    
    # Cast negative distances as 10000 (non-continuity)
    #    - This forces us to only look to the right
    delta_x = np.where(delta_x<0, 10000, delta_x)
    
    # Find the y distances between cp and all cps
    delta_y = np.abs((cps[..., 1] - cp[1]))
    
    # Find the combined distance after multiplying by new
    # coordinate system distances in each respective direction
    delta = (delta_x*x_step)+(delta_y*y_step)
    
    # Find the cp in the list of cps that is closest to the original cp
    # using the deltas as calculated with the new coordinate system rules
    min_index = np.argmin(delta)
    
    # If the nearest cp is VERY FAR AWAY (>10000) or 
    # if the nearest cp is more than max_dy away in the vertical direction
    # than we return None... else we return the index of the closest cp
    # to the original cp from the list of cps.
    if np.abs(cps[min_index][1]-cp[1])>max_dy or delta[min_index]>10000:
        return None
    else:
        return min_index
    

def get_bb_lbl_map_1(path_to_resized, preds, c_thresh, make_flexible=True, upper_cp=False, max_skew_angle=5):
    """ Algorithmic Approach 1 For Retrieving the BBOX Label Map 
    
    Args:
        path_to_resized (str): Path to the resized slide image saved to disk
        preds (AI Platform Predictions): Raw predictions from AI Platform
        c_thresh (float, optional): Confidence threshold to filter predictions upon
        make_flexible (bool, optional): Whether to increase row height 
            allowing for more flexibility
        upper_cp (bool, optional): Whether to use upper quarter instead of midpoint as
            the 'centre-point'
        max_skew_angle (int, optional): Maximum value for slide to be considered skewed
    
    Returns:
        A dictionary mapping rows (labelled with capital letters starting at `A`) 
        to a numpy array where each row in the array is the bbox coordinates for
        a tissue sample (sorted from left to right)
    """
    
    # Get dimensions of resized image from filename
    dimensions = np.array(path_to_resized[:-4].rsplit("_", 2)[1:], np.int32)
    
    # Filter confidences
    confidences_raw = preds[0]["confidences"]
    confidences = np.array([x for x in confidences_raw if x>c_thresh])
    
    # Capture bboxes, confidences and centre points above the threshold
    bboxes = np.zeros((confidences.size, 4), dtype=np.float32)
    centre_points = np.zeros((confidences.size, 2), dtype=np.float32)
    for i, bbox in enumerate(preds[0]["bboxes"]):
        if confidences_raw[i]>=c_thresh:
            bboxes[i, :2] = np.array(bbox)[[0,2]]
            bboxes[i, 2:] = np.array(bbox)[[1,3]]
            if not upper_cp:
                centre_points[i, :] = np.array([bbox[1]-(bbox[1]-bbox[0])/2, bbox[3]-(bbox[3]-bbox[2])/2])
            else:
                centre_points[i, :] = np.array([bbox[0]+(bbox[1]-bbox[0])/3, bbox[2]+(bbox[3]-bbox[2])/3])
    
    # Calculate skew of slide
    (cx, cy), (w,h), theta = cv2.minAreaRect(np.array(centre_points))
    skew_angle = 90%abs(theta)
    if skew_angle > max_skew_angle:
        logging.info("\n\n======================================================") 
        logging.info("... Slide Is Skewed By At Least %s Degrees.Algorithmic Approach May Fail! ...",\
              str(skew_angle))  
        logging.info("=============================================================\n\n")
    
    # Get bbox rows
    bb_rows = compute_bb_rows_1(bboxes, centre_points, make_flexible)
    
    # Map bboxes to letter of the row for appropriate naming
    bb_lbl_map = {chr(65+i):bb for i, bb in enumerate(bb_rows)}
    return bb_lbl_map


def get_bb_lbl_map_2(path_to_resized, preds_confidences, preds_bbox , C_THRESH , dy_frac, tl_find_mult, do_nms ):
    """ Algorithmic Approach 2 For Retrieving the BBOX Label Map 
    
    Args:
        path_to_resized (str): Path to the resized slide image saved to disk
        preds (AI Platform Predictions): Raw predictions from AI Platform
        dy_frac (float, optional): Fraction of bbox height to allow in vertical direction
        tl_find_mult (int, optional): Percentage to weight y dim vs x dim in tl corner search
    
    Returns:
        A dictionary mapping rows (labelled with capital letters starting at `A`) 
        to a numpy array where each row in the array is the bbox coordinates for
        a tissue sample (sorted from left to right)
    """    
    
    # Setup
    #dimensions = np.array(path_to_resized[:-4].rsplit("_", 2)[1:], np.int32)
    dimensions = np.array(path_to_resized.rsplit(".", 1)[0].rsplit("_", 2)[1:], np.int32) 
    #dimensions = np.array(path_to_resized.rsplit(".", 2)[0].rsplit("_", 2)[1:], np.int32)
    #raw_confidences = preds[0]["confidences"]
    raw_confidences = preds_confidences
    confidences = np.array([x for x in raw_confidences if x>C_THRESH])
    bboxes = np.zeros((confidences.size, 4), dtype=np.float32)
    centre_points = np.zeros((confidences.size, 2), dtype=np.float32)

    for i, bbox in enumerate(preds_bbox):
        if raw_confidences[i] > C_THRESH:
            bboxes[i, :2] = np.array(bbox)[[0,2]]
            bboxes[i, 2:] = np.array(bbox)[[1,3]]
            centre_points[i, :] = np.array([bbox[0]+(bbox[1]-bbox[0])/2, bbox[2]+(bbox[3]-bbox[2])/2])

    if do_nms:
        bboxes, centre_points = bbox_suppression(bboxes, centre_points)
    
    # Initialization
    bb_lbl_map = {}
    max_dy = np.mean(np.diff(bboxes[:, [1,3]]))*dy_frac
    chr_idx = 65

    # The magic
    while len(bboxes)>0:
        # Get the new top_left bbox and remove it from cps and bbs
        tl_point = centre_points[np.argmin(bboxes[:, 1]*tl_find_mult+bboxes[:, 0])][:2]-0.01

        # First pass on repetition
        closest_idx = closest_cp(tl_point, centre_points, max_dy*2)

        current_cp = centre_points[closest_idx]
        bb_lbl_map[chr(chr_idx)] = [bboxes[closest_idx],]
        
        bboxes=np.delete(bboxes, closest_idx, axis=0)
        centre_points=np.delete(centre_points, closest_idx, axis=0)

        while len(bboxes)>0:
            closest_idx = closest_cp(current_cp, centre_points, max_dy)
            if closest_idx is None:
                break
            current_cp = centre_points[closest_idx]
            bb_lbl_map[chr(chr_idx)].append(bboxes[closest_idx])
            bboxes=np.delete(bboxes, closest_idx, axis=0)
            centre_points=np.delete(centre_points, closest_idx, axis=0)

        # Next Row Character
        chr_idx+=1
        
    bb_lbl_map = {k:np.stack(v) for k,v in bb_lbl_map.items()}
    return bb_lbl_map


def plot_image(bb_lbl_map, path_to_resized, opacity, cmap, plot_row_boundaries, plot_box_midline, upper_cp):
    """ Plot Bounding Boxes By Row 
    
    Args:
        bb_lbl_map (dictionary): Mapping from row to numpy array containing bboxes
        path_to_resized (str): Path to downscaled image on disk
        opacity (float, optional): Opacity of the overlayed square for display
        cmap (string, optional): SNS Colourmap to be used
        plot_row_boundaries (bool, optional): Whether to show the edges of the rows
        plot_box_midline (bool, optional): Whether to show the centre-line of each box
        upper_cp (bool, optional): Whether to see the cetnre-point as the top quartile.

    Returns:
        None; An image will be plotted
    """
    def __get_scaled_bbox(bbox, dims):
        """ Scale the bounding box from float to the approrpiate integer dimensions """
        return int(bbox[0]*dims[1]), int(bbox[1]*dims[0]), int(bbox[2]*dims[1]), int(bbox[3]*dims[0])
    
    # Initialize arrays and colours
    COLORS = [(int(x[0]*255), int(x[1]*255), int(x[2]*255)) for x in sns.color_palette(cmap, len(bb_lbl_map))]
    random.shuffle(COLORS)
    FONT = cv2.FONT_HERSHEY_SIMPLEX; FONT_SCALE = 0.7; FONT_THICKNESS = 2; FONT_LINE_TYPE = cv2.LINE_AA;
    #path_to_resized = "26106_TMA22_RGB_1120_835.jpeg" RGB to be used
    # Open the image and grab the dimensions
    path_to_resized = path_to_resized.replace("Original", "RGB").replace("png", "jpeg")
    
    img = np.asarray(Image.open(path_to_resized)).copy()
    dims = img.shape
      
    # Go row by row and plot the required shapes/text/lines
    plt.figure(figsize=(18,18))
    for i, (row_ltr, bb_list) in enumerate(bb_lbl_map.items()):
        if plot_row_boundaries:
            ymin = bb_list[np.argmin(bb_list[:, 1])][1]*dims[0]
            ymax = ((bb_list[np.argmax(bb_list[:, 3])][3]+bb_list[np.argmin(bb_list[:, 3])][3])/2)*dims[0]
            plt.hlines(ymin, 0, dims[1], colors=tuple([max((c-40)/255, 0.0) for c in COLORS[i]]+[0.45,]), linestyles='solid', linewidth=5)
            plt.hlines(ymax, 0, dims[1], colors=tuple([max((c-40)/255, 0.0) for c in COLORS[i]]+[0.45,]), linestyles='solid', linewidth=5)
        for j, bbox in enumerate(bb_list):
            
            bbox = __get_scaled_bbox(bbox, dims)
            rect = np.uint8(np.ones((bbox[3]-bbox[1], bbox[2]-bbox[0], 3))*COLORS[i])
            cv2.rectangle(img, (bbox[0], bbox[1]), (bbox[2], bbox[3]), tuple([max((c-40), 0) for c in COLORS[i]]), FONT_THICKNESS)
            
            overlay = cv2.addWeighted(img[bbox[1]:bbox[3],bbox[0]:bbox[2],:], 1-opacity, rect, opacity, 1.0)
            img[bbox[1]:bbox[3],bbox[0]:bbox[2],:] = overlay
            
            if plot_box_midline:
                if not upper_cp:                    
                    plt.hlines(bbox[1]+(bbox[3]-bbox[1])//2, bbox[0], bbox[2], colors=tuple([max((c-40)/255, 0) for c in COLORS[i]]+[0.75,]), linestyles='dashed', linewidth=2)
                else:
                    plt.hlines(bbox[1]+(bbox[3]-bbox[1])//3, bbox[0], bbox[2], colors=tuple([max((c-40)/255, 0) for c in COLORS[i]]+[0.75,]), linestyles='dashed', linewidth=2)
            
            txt_lbl = f"{row_ltr}{j+1}"
            text_width, text_height = cv2.getTextSize(txt_lbl, FONT, FONT_SCALE, FONT_THICKNESS)[0]
            cv2.putText(img, txt_lbl, (bbox[0]+((bbox[2]-bbox[0])//2)-(text_width//2), (bbox[1]+((bbox[3]-bbox[1])//2)+(text_height//2))),
                        FONT, FONT_SCALE, [max((c-70), 0) for c in COLORS[i]], FONT_THICKNESS, FONT_LINE_TYPE)

    plt.imshow(img)
    plt.title(path_to_resized)
    plt.axis(False)
    #plt.savefig(fname = os.getcwd())  #to be removed just for testing
    return plt

###TMA classification model

def predict_cloud_model(    
    items,
    project: str,
    endpoint_id: str, 
    C_THRESH,
    config,
    job_name,
    location: str = "us-central1",
    api_endpoint: str = "us-central1-prediction-aiplatform.googleapis.com",
    instance_dict: dict = None
    
):
    """ Get predictions from Google AIPlatform Endpoint """
    client_options = {"api_endpoint": api_endpoint}
    
    # Initialize client that will be used to create and send requests.
    # This client only needs to be created once, and can be reused for multiple requests.
    client = aiplatform.gapic.PredictionServiceClient(client_options=client_options)
    
    gcs_path, new_dims, name , local_svs_path = items
    
    # filename = f"{name}_Original_{new_dims[0]}_{new_dims[1]}.png"
    # updated on 08-03-2022
    filename = f"{name}_Gray_{new_dims[0]}_{new_dims[1]}.png"
    with open(filename, "rb") as f:
        file_content = f.read()
    # The format of each instance should conform to thef deployed model's prediction input schema.
    encoded_content = base64.b64encode(file_content).decode("utf-8")
    
    if not instance_dict:
        instance_dict = {"content": encoded_content}
    
    instance = json_format.ParseDict(instance_dict, Value())
    instances = [instance]
    # See gs://google-cloud-aiplatform/schema/predict/params/image_object_detection_1.0.0.yaml for the format of the parameters.
    parameters_dict = {"confidenceThreshold": C_THRESH, "maxPredictions": 200}
    parameters = json_format.ParseDict(parameters_dict, Value())
    endpoint = client.endpoint_path(
        project=project, location=location, endpoint=endpoint_id
    )
    response = client.predict(
        endpoint=endpoint, instances=instances, parameters=parameters,
    )
    logging.info("response")
    deployed_model_id = response.deployed_model_id
    logging.info(" deployed_model_id:  "+ str(deployed_model_id))
    # See gs://google-cloud-aiplatform/schema/predict/prediction/image_object_detection.yaml for the format of the predictions.
    preds = response.predictions
    preds, label = {k:v for k,v in preds[0].items()}, preds[0]["displayNames"][np.argmax(preds[0]["confidences"])]   
    if label == "":
        logging.exception("Classification model response in blank ")        
        yield pvalue.TaggedOutput("classfication_error",(gcs_path.split(',')[0],config["function_config"]["Insert_Error_BQ"]["dataset_id"],config["function_config"]["Insert_Error_BQ"]["table_id"],job_name,'TMA classifier - Pred',"Classification model response is blank"))
    else:
         yield pvalue.TaggedOutput("classfication_svs",(gcs_path,label,filename,local_svs_path))
    
    
def predict_cloud_model_2(
    project: str,
    endpoint_id: str,
    encoded_content,
    location: str = "us-central1",
    api_endpoint: str = "us-central1-prediction-aiplatform.googleapis.com",
    instance_dict: dict = None,
):
    """ Get predictions from Google AIPlatform Endpoint """
    client_options = {"api_endpoint": api_endpoint}
    
    # Initialize client that will be used to create and send requests.
    # This client only needs to be created once, and can be reused for multiple requests.
    client = aiplatform.gapic.PredictionServiceClient(client_options=client_options)
    
    '''with open(filename, "rb") as f:
        file_content = f.read()'''

    # The format of each instance should conform to the deployed model's prediction input schema.
    #encoded_content = base64.b64encode(file_content).decode("utf-8")
    
    if not instance_dict:
        instance_dict = {"input_1": encoded_content}
    
    instance = json_format.ParseDict(instance_dict, Value())
    instances = [instance]
    
    # See gs://google-cloud-aiplatform/schema/predict/params/image_object_detection_1.0.0.yaml for the format of the parameters.
    C_THRESH=0.025
    parameters_dict = {"confidenceThreshold": C_THRESH, "maxPredictions": 200}
    parameters = json_format.ParseDict(parameters_dict, Value())
    endpoint = client.endpoint_path(
        project=project, location=location, endpoint=endpoint_id
    )
    response = client.predict(
        endpoint=endpoint, instances=instances, parameters=parameters,
    )
    logging.info("response")
    deployed_model_id = response.deployed_model_id
    logging.info(" deployed_model_id: %s", str(deployed_model_id))
    # See gs://google-cloud-aiplatform/schema/predict/prediction/image_object_detection.yaml for the format of the predictions.
    predictions = response.predictions
    return predictions

@tf.function
def rgb2hsd(img):
    lower_clip = 1e-2
    eps = tf.constant(1e-7)
    log10 = tf.math.log(tf.constant(10.0))
    img = tf.image.convert_image_dtype(img, tf.float32)
    img = tf.where(img<lower_clip, tf.multiply(lower_clip, tf.ones_like(img)), img)
    OD = tf.divide(tf.multiply(tf.constant(-1.0), tf.math.log(img)), log10)
    D = tf.reduce_mean(OD, axis=-1)
    D = tf.where(D==0.0, tf.multiply(eps, tf.ones_like(D)), D)
    cx = tf.subtract(tf.divide(OD[:,:,0], D), tf.constant(1.0))
    cx = tf.where(cx==0.0, tf.multiply(eps, tf.ones_like(cx)), cx)
    cy = tf.divide(tf.subtract(OD[:,:,1], OD[:,:,2]), tf.multiply(tf.sqrt(tf.constant(3.0)), D))
    D = tf.expand_dims(D, -1)
    cx = tf.expand_dims(cx, -1)
    cy = tf.expand_dims(cy, -1)
    img = tf.concat((D,cx,cy), -1)
    return img


def crop_image(idx, offset_height=0, offset_width=0, size=1024):
    image = full_dataset['image'][idx]
    image = tf.convert_to_tensor(image)
    image = tf.image.convert_image_dtype(image, tf.float32)
    image = tf.image.crop_to_bounding_box(
        image, 
        offset_height=offset_height, 
        offset_width=offset_width, 
        target_height=size,
        target_width=size
      )
    return image


def clean_image(image, sigma=3, t=0.9):
    blur = skimage.color.rgb2gray(image)
    blur = skimage.filters.gaussian(blur, sigma=sigma)
    mask = blur < t
    
    sel = np.ones_like(image)*255
    sel[mask] = image[mask]
    
    image = tf.convert_to_tensor(sel)
   
    return image

def preprocess_cd228(img, c_crop=0.85, resize_to=(2048,2048)):
    #print("img.shape", img.shape)
    img = clean_image(img)
    img = rgb2hsd(img)
    img = tf.image.central_crop(img, c_crop)
    img = tf.image.resize(img, resize_to)
    return img


def preprocess_slc1a5(img, c_crop=0.85, resize_to=(2048,2048)):
    img = clean_image(img, 2, 0.93)
    img = tf.image.central_crop(img, c_crop)
    img = tf.cast(tf.image.resize(img, resize_to), dtype=tf.uint8)  
    return img/255

def postprocess_slc1a5(pred_map, multiplier_map={"s":1.0, "i":3.0, "f":4.0}):
    for k,v in  multiplier_map.items():
        pred_map[k] = round(pred_map[k]*v, 3)
    return pred_map

def predict_cloud_model_patho(
    project: str,
    endpoint_id: str,
    filename: str,
    location: str = "us-central1",
    api_endpoint: str = "us-central1-prediction-aiplatform.googleapis.com",
    instance_dict: dict = None,
):
    """ Get predictions from Google AIPlatform Endpoint """
    client_options = {"api_endpoint": api_endpoint}
    
    # Initialize client that will be used to create and send requests.
    # This client only needs to be created once, and can be reused for multiple requests.
    client = aiplatform.gapic.PredictionServiceClient(client_options=client_options)
    
    with open(filename, "rb") as f:
        file_content = f.read()
        
    # The format of each instance should conform to the deployed model's prediction input schema.
    encoded_content = base64.urlsafe_b64encode(file_content).decode("utf-8")
    
    if not instance_dict:
        instance_dict = {"image_bytes": encoded_content}
    
    instance = json_format.ParseDict(instance_dict, Value())
    instances = [instance]
    #instances = [{"image_bytes": {"b64": encoded_content}}]
    logging.info("***Estimated size encoding:  %s MB ",str(format(sys.getsizeof(encoded_content) / 1024**2)))
    C_THRESH=0.025
    parameters_dict = {"confidenceThreshold": C_THRESH, "maxPredictions": 200}
    parameters = json_format.ParseDict(parameters_dict, Value())
    endpoint = client.endpoint_path(
        project=project, location=location, endpoint=endpoint_id
    )
    
    response = client.predict(
        endpoint=endpoint, instances=instances, parameters=parameters,
    )
    logging.info("response")
    deployed_model_id = response.deployed_model_id
    logging.info(" deployed_model_id:  "+ str(deployed_model_id))        
    # See gs://google-cloud-aiplatform/schema/predict/prediction/image_object_detection.yaml for the format of the predictions.
    predictions = response.predictions
    return predictions


def check_if_none(list_of_elem):
    """ Check if elements in list are None """
    result = False
    for elem in list_of_elem:
        if elem is not None:
            return True
    return result



def load_biomax_csv(items,dataset_id,table_id,config,job_name):
    #dataset_id = "seagen_quantiphi"
    #bq table name
    #table_id = "biomax_metadata"
    # loading the file from gcs to local folder
    gcs_path , biomax_excel_local = items[1]
    logging.info("biomax_excel_local "+ str(biomax_excel_local))
    # loading the excel workbook using openpyxl load workbook function
    
    ##check file type
    file_type = biomax_excel_local.rsplit('.',1)[-1]
    assert file_type == "xlsx" , "Only xlsx files are allowed for Biomax"
    wb = load_workbook (biomax_excel_local)
    
    # creating worksheet object
    ws = wb.worksheets[0]
    # total no of rows = B8*B9
    # 
    try:
        
        num_rows = int(ws["B8"].value)*int(ws["B9"].value)
        # total columns = 11, bq will have same no of columns
        num_cols = 11   
        ##name B1 = BL241b
        Name = ws["B1"].value

        ##verify header row
        ##check header count ; should not be less 11
        header = [cell.value for cell in ws[11]]
        assert len(header) >= 11 , "Number of cloumns in Biomax excel are less than 11"

        BQ_list=[]  # List of dict
        for i in range(12 , num_rows+12):     # loop for rows
            row_data=[] # list to read one row of data
            for j in range(1, num_cols + 1):   # loop for columns
                cell_obj = ws.cell(row = i, column = j)
                #create list with 1 row of data
                row_data.append(cell_obj.value)
            #check none row
            print(row_data)
            result = check_if_none(row_data)
            #create dict
            if (result):
                    try:
                         BQ_list.append({"Position": f"{row_data[0][0]}{int(row_data[0][1:]):02}", "Number": row_data[1] ,                          "Age": row_data[2] , "Sex": row_data[3],"Organ_AnatomicSite":row_data[4],
                                "Pathology_diagnosis":row_data[5],
                                "TNM":row_data[6], "Grade":row_data[7], "Stage":row_data[8], "Type":row_data[9],
                                "Tissue_ID":row_data[10]                       
                           })
                    except TypeError:
                            logging.info("Position column is blank")

        # calling bq API    
        client = bigquery.Client()
        dataset_ref = client.dataset(dataset_id)
        table_ref = dataset_ref.table(table_id)
        table = client.get_table(table_ref)  # API call

        #verfiy if sheet Name exists in table
        query_job = client.query("""
            SELECT TMA_Name as SName
            FROM `ihc-qc-sandbox.seagen_quantiphi.Biomax_Metadata` 
            """)

        #inserting data in bq
        
        rows_to_insert = [{"TMA_Name": Name,
                               "Name": BQ_list,
                               "job_name":job_name
                              }]
        errors = client.insert_rows_json(table, rows_to_insert)
        
    except Exception as e:
        logging.exception("incorrect field values-B8,B9 for Biomax excel "+str(e))
        Insert_Error_BQ((gcs_path,config["function_config"]["Insert_Error_BQ"]["dataset_id"],config["function_config"]["Insert_Error_BQ"]["table_id"],job_name, 'load_biomax_csv',str(e)))
        pass
        
        
    
def Insert_Error_BQ (items):
    # calling bq API  
    image_path,dataset_id,table_id,job_name,source, error_reason  =items
    #image_path = items
    client = bigquery.Client()
    dataset_ref = client.dataset(dataset_id)
    table_ref = dataset_ref.table(table_id)
    table = client.get_table(table_ref)  # API call
    
    rows_to_insert = [{"Job_Name": job_name,
                       "Image_Path": image_path,
                       "error_source":source,
                       "exception":error_reason
                          }]
    errors = client.insert_rows_json(table, rows_to_insert)
        
def plot_image_quad(bb_lbl_map, path_to_resized, opacity=0.5, cmap="husl", plot_row_boundaries=True, plot_box_midline=False, upper_cp=False):
    """ Plot Bounding Boxes By Row 
    
    Args:
        bb_lbl_map (dictionary): Mapping from row to numpy array containing bboxes
        path_to_resized (str): Path to downscaled image on disk
        opacity (float, optional): Opacity of the overlayed square for display
        cmap (string, optional): SNS Colourmap to be used
        plot_row_boundaries (bool, optional): Whether to show the edges of the rows
        plot_box_midline (bool, optional): Whether to show the centre-line of each box
        upper_cp (bool, optional): Whether to see the cetnre-point as the top quartile.

    Returns:
        None; An image will be plotted
    """
    def __get_scaled_bbox(bbox, dims):
        """ Scale the bounding box from float to the approrpiate integer dimensions """
        return int(bbox[0]*dims[1]), int(bbox[1]*dims[0]), int(bbox[2]*dims[1]), int(bbox[3]*dims[0])
    
    # Initialize arrays and colours
    COLORS = [(int(x[0]*255), int(x[1]*255), int(x[2]*255)) for x in sns.color_palette(cmap, len(bb_lbl_map))]
    
    #print('COLORS:\n', COLORS, print('COLORS.len ', len(COLORS))
    
    random.shuffle(COLORS)
    FONT = cv2.FONT_HERSHEY_SIMPLEX; FONT_SCALE = .7; FONT_THICKNESS = 2; FONT_LINE_TYPE = cv2.LINE_AA;
    
    # Open the image and grab the dimensions
    #img = np.asarray(Image.open(path_to_resized.replace(".png", "__RGB_VERSION.png"))).copy()
    img = np.asarray(Image.open(path_to_resized)).copy()
    dims = img.shape
    
    quad_list = ['A','B','C','D']
    target_col_len = get_most_common_max_row_len(bb_lbl_map)
    target_row_len = len(bb_lbl_map)
        
    # generate row label list based on 1/2 the num of rows
    chr_idx = 65 #+int(target_row_len/2)-1
    quad_row_list = []
    for ii in range(int(target_row_len/2)):
        quad_row_list.append(chr(chr_idx))
        chr_idx += 1
            
    #print('quad_row_list: ', quad_row_list)
    #print('target_row_len: ', target_row_len, '\ttarget_col_len: ', target_col_len)
    
    # Go row by row and plot the required shapes/text/lines
    plt.figure(figsize=(18,18))
    for i, (row_ltr, bb_list) in enumerate(bb_lbl_map.items()):
        if plot_row_boundaries:
            ymin = bb_list[np.argmin(bb_list[:, 1])][1]*dims[0]
            ymax = ((bb_list[np.argmax(bb_list[:, 3])][3]+bb_list[np.argmin(bb_list[:, 3])][3])/2)*dims[0]
            plt.hlines(ymin, 0, dims[1], colors=tuple([max((c-40)/255, 0.0) for c in COLORS[i]]+[0.45,]), linestyles='solid', linewidth=5)
            plt.hlines(ymax, 0, dims[1], colors=tuple([max((c-40)/255, 0.0) for c in COLORS[i]]+[0.45,]), linestyles='solid', linewidth=5)
        for j, bbox in enumerate(bb_list):
            
            bbox = __get_scaled_bbox(bbox, dims)
            rect = np.uint8(np.ones((bbox[3]-bbox[1], bbox[2]-bbox[0], 3))*COLORS[i])
            cv2.rectangle(img, (bbox[0], bbox[1]), (bbox[2], bbox[3]), tuple([max((c-40), 0) for c in COLORS[i]]), FONT_THICKNESS)
            overlay = cv2.addWeighted(img[bbox[1]:bbox[3],bbox[0]:bbox[2],:], 1-opacity, rect, opacity, 1.0)
            img[bbox[1]:bbox[3],bbox[0]:bbox[2],:] = overlay
            
            if plot_box_midline:
                if not upper_cp:                    
                    plt.hlines(bbox[1]+(bbox[3]-bbox[1])//2, bbox[0], bbox[2], colors=tuple([max((c-40)/255, 0) for c in COLORS[i]]+[0.75,]), linestyles='dashed', linewidth=2)
                else:
                    plt.hlines(bbox[1]+(bbox[3]-bbox[1])//3, bbox[0], bbox[2], colors=tuple([max((c-40)/255, 0) for c in COLORS[i]]+[0.75,]), linestyles='dashed', linewidth=2)
            
            if i+1 <= target_row_len/2 and j < target_col_len/2:  # QUAD A
                row_prefix = quad_list[0]+row_ltr
                quad_num = j
                    
            elif i+1 <= target_row_len/2 and j >= target_col_len/2:  # QUAD B
                row_prefix = quad_list[1]+row_ltr
                quad_num = j%(target_col_len/2)
                    
            elif i+1 > target_row_len/2 and j < target_col_len/2:  # QUAD C
                quad_num = j
                quad_row = int(i%(target_row_len/2))
                row_prefix = quad_list[2] + quad_row_list[quad_row]
                    
            elif i+1 > target_row_len/2 and j >= target_col_len/2:  # QUAD D
                quad_num = j%(target_col_len/2)
                quad_row = int(i%(target_row_len/2))
                row_prefix = quad_list[3] + quad_row_list[quad_row]
                
            quad_num += 1
            #print('row_prefix: ', row_prefix, '\tquad_num: ', quad_num)
                
            #txt_lbl = f"{row_ltr}{j+1}"
            txt_lbl = row_prefix+str(int(quad_num))
            
            text_width, text_height = cv2.getTextSize(txt_lbl, FONT, FONT_SCALE, FONT_THICKNESS)[0]
            cv2.putText(img, txt_lbl, (bbox[0]+((bbox[2]-bbox[0])//2)-(text_width//2), (bbox[1]+((bbox[3]-bbox[1])//2)+(text_height//2))),
                        FONT, FONT_SCALE, [max((c-70), 0) for c in COLORS[i]], FONT_THICKNESS, FONT_LINE_TYPE)

    plt.imshow(img)
    #plt.title(path_to_resized.rsplit("/", 1)[1], fontweight="bold")
    plt.title(path_to_resized, fontweight="bold")
    plt.axis(False)
    plt.savefig('map_'+path_to_resized)
    plt.show()
    
def get_most_common_max_row_len(map_dict):
    row_lens = []
        
    # get the target elements per row - based on the most common 
    for tma_row_key in map_dict.keys():
        row_lens.append(len(map_dict[tma_row_key]))
    set_row_len = max(set(row_lens), key = row_lens.count)
    #print('set_row_len: ', set_row_len)
    return set_row_len

    
def get_avg_box_len(map_dict_keys):
    box_len_sum = 0.0
    for column, box in enumerate(map_dict_keys):
        box_len_sum += box[2] - box[0] # x1 - x0
    avg_box_len = box_len_sum/len(map_dict_keys)
    return avg_box_len

    
def tma_core_correction(fname, bb_lbl_map, left_justify=False, fill_last_row=False):
        
    bbox_diff_max = .02 #0.29 better for 100x downscale  #should calculate based on % of avg. width of boxes

    ## For now, we will set "correction" parameters here
    ## - using fname to determnine the type of TMA

    if 'champions' in fname.lower():
        bbox_diff_max = .012 #.012
        fill_last_row = True
        left_justify = True
    elif 'biomax_quad' in fname.lower():
        bbox_diff_max = .035
             
    bb_lbl_map_curr = bb_lbl_map
        
    left_justify_coords = []

    set_row_len = get_most_common_max_row_len(bb_lbl_map_curr)
        
    # get the number of rows
    num_rows = len(bb_lbl_map.keys())
    #print('num_rows: ', num_rows)
        
    for row, tma_row_key in enumerate(bb_lbl_map.keys()):
        prev_box = []
            
        # after processing each row, update to get the most common row max
        set_row_len = get_most_common_max_row_len(bb_lbl_map_curr)
            
        # get the avg length for boxes in the row - used for new box generation
        avg_row_box_len = get_avg_box_len(bb_lbl_map[tma_row_key])
            
        for column, box in enumerate(bb_lbl_map[tma_row_key]):
            curr_box = box
            #print('curr_box\n:', curr_box)
            #print('curr_box\n:', prev_box)
                
            # only happens for the first core
            if len(left_justify_coords) == 0:
                left_justify_coords = curr_box
                    
            if left_justify and len(prev_box) == 0 and ((curr_box[0] - left_justify_coords[0]) > bbox_diff_max):
                #print('..found missing LEFT JUSTIFY bbox: key: ', tma_row_key, '\t column: ', column, '\t diff: ', curr_box[0] - left_justify_coords[0])
                
                new_box = [left_justify_coords[0], curr_box[1], left_justify_coords[2], curr_box[3]]
                #print('...appending to dict: ', new_box)
                box_list = list(bb_lbl_map_curr[tma_row_key])
                box_list.append(new_box)
                box_list.sort(key = lambda x: x[0])
                bb_lbl_map_curr[tma_row_key] = np.array(box_list)
                    
                prev_box = new_box
                
            # if curr_box is not next to prev box (prev_box x1 is "near" curr_box x0)
            while (len(prev_box) > 0) and ((curr_box[0] - prev_box[2]) > bbox_diff_max):
                    
                #print('..found missing bbox: key: ', tma_row_key, '\t column: ', column, '\t diff: ', curr_box[0] - prev_box[2])
                
                # calc # of boxes missing, generate coords for each, insert into key/column bb_lbl_map
                #new_box = [prev_box[2], prev_box[1], prev_box[2]+(prev_box[2]-prev_box[0]), prev_box[3]]  # here we are setting the width to that of the prev_box
                new_box = [prev_box[2], prev_box[1], prev_box[2]+(avg_row_box_len), prev_box[3]]  # here we are setting the width to AVG fo boxes in the row
                    
                #print('...appending to dict: ', new_box)
                box_list = list(bb_lbl_map_curr[tma_row_key])
                box_list.append(new_box)
                box_list.sort(key = lambda x: x[0])
                bb_lbl_map_curr[tma_row_key] = np.array(box_list)
                    
                prev_box = new_box
                
            prev_box = curr_box
            
        if not fill_last_row and (row == num_rows-1):
            print(' fill_last_row is False and on last row: ', row, '...skippiing...')
        else:
            # if more cores should be added at the end of the row, this is where we do it.
            for ii in range(set_row_len - len(bb_lbl_map_curr[tma_row_key])):

                #print('..found missing bbox AT END OF ROW: key: ', tma_row_key, '\t ii: ', ii)

                prev_boxes = list(bb_lbl_map_curr[tma_row_key])
                prev_box = prev_boxes[-1]
                
                # calc # of boxes missing, generate coords for each, insert into key/column bb_lbl_map
                #new_box = [prev_box[2], prev_box[1], prev_box[2]+(prev_box[2]-prev_box[0]), prev_box[3]]
                new_box = [prev_box[2], prev_box[1], prev_box[2]+(avg_row_box_len), prev_box[3]] 
                for jj in range(len(new_box)):
                    if new_box[jj] > 1:
                        new_box[jj] = 1.0

                #print('...appending to dict: ', new_box)
                if (new_box[2]-new_box[0]) > 0.0:
                    prev_boxes.append(new_box)
                    prev_boxes.sort(key = lambda x: x[0])
                    bb_lbl_map_curr[tma_row_key] = np.array(prev_boxes)
                 
    return bb_lbl_map_curr 
